from flask import Flask, render_template, redirect, url_for, send_from_directory, flash, request , jsonify, session
import firebase_admin
from firebase_admin import credentials, storage, firestore, db
from datetime import datetime, timedelta
import os
import io
import zipfile
import logging
import openpyxl  # Libreria per gestire Excel
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import shutil
from forms import LoginForm  # Importa il form di login
import json


import xlsxwriter

# Carica le variabili d'ambiente dal file .env
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')  

firebase_key_json = os.getenv('FIREBASE_KEY_JSON')

app.debug = False

# Utente e password caricati dal file .env
USER_NAME = os.getenv('USER_NAME')  # Nome utente personalizzato
USER_PASSWORD = os.getenv('USER_PASSWORD')  # Password definita nel file .env

print(f"User name: {USER_NAME}")
print(f"User password: {USER_PASSWORD}")

# Inizializza il logger
logging.basicConfig(level=logging.DEBUG if app.debug else logging.INFO)

# Inizializza Firebase
if not firebase_admin._apps:
    cred = credentials.Certificate(json.loads(firebase_key_json))
    firebase_admin.initialize_app(cred, {
        'storageBucket': os.getenv('FIREBASE_STORAGE_BUCKET'),
        'databaseURL': os.getenv('FIREBASE_DB_URL')  # Aggiungi l'URL del database in tempo reale
    })

bucket = storage.bucket()
db_ref = db.reference("/Attivita/Utenti")

# Directory temporanea per i file
TEMP_DIR = os.path.join(os.getcwd(), "temp")
os.makedirs(TEMP_DIR, exist_ok=True)

def clean_temp_directory():
    try:
        shutil.rmtree(TEMP_DIR)  # Rimuovi la directory temporanea e tutto il suo contenuto
        os.makedirs(TEMP_DIR, exist_ok=True)  # Ricrea la directory vuota
        logging.info("Cartella temp pulita con successo.")
    except Exception as e:
        logging.error(f"Errore durante la pulizia della cartella temp: {e}")

clean_temp_directory()


# Rotta per il login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        # Verifica se l'username e la password corrispondono a quelli nel file .env
        if username == USER_NAME and password == USER_PASSWORD:
            session['user'] = username
            return redirect(url_for('index'))
        else:
            flash('Credenziali non valide. Riprova.')

    return render_template('login.html')

# Rotta per l'homepage dopo il login
@app.route('/')
def index():
    if 'user' in session:
        clean_temp_directory()
        return render_template('index.html')
    else:
        return redirect(url_for('login'))

# Rotta per il logout
@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('Sei stato disconnesso.')
    return redirect(url_for('login'))


@app.route('/gestione_cantieri')
def gestione_cantieri():
    return render_template('gestione_cantieri.html')

@app.route('/gestione_operai')
def gestione_operai():
    return render_template('gestione_operai.html')

@app.route('/gestione_excel', methods=['GET', 'POST'])
def gestione_excel():

    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    logging.info("Accesso alla rotta gestione_excel")
    if request.method == 'POST':
        logging.info("Metodo POST chiamato")
        data_inizio = request.form['start_date']
        data_fine = request.form['end_date']
        action = request.form['action']  # Può essere 'contabilita', 'completo', 'buste'

        logging.info(f"Data Inizio: {data_inizio}, Data Fine: {data_fine}, Azione: {action}")

        attivita_list = fetch_attivita_from_firebase(data_inizio, data_fine)

        if action == 'contabilita':
            file_path = generate_excel_contabilita(attivita_list, data_inizio, data_fine)
        elif action == 'completo':
            file_path = generate_excel_completo(attivita_list, data_inizio, data_fine)
        elif action == 'buste':
            file_path = generate_excel_buste(attivita_list, data_inizio, data_fine)

        # Restituisci il file Excel generato come risposta
        if file_path:
            return send_from_directory(TEMP_DIR, file_path, as_attachment=True)


    return render_template('gestione_excel.html')


def fetch_attivita_from_firebase(data_inizio, data_fine):
    ref = db.reference('/Attivita/Utenti')
    snapshot = ref.get()
    attivita_list = []

    for utente_id, utente_data in snapshot.items():
        for attivita_id, attivita in utente_data.items():
            data_attivita = attivita.get('data')
            if is_date_in_range(data_attivita, data_inizio, data_fine):
                attivita_list.append({
                    'data': attivita.get('data'),
                    'cantiere': attivita.get('cantiere'),
                    'operaio': attivita.get('operaio'),
                    'ore': attivita.get('ore'),
                    'lavorazione': attivita.get('lavorazione'),
                    'pioggia_vento': attivita.get('pioggia_vento'),
                    'ferie_permesso': attivita.get('ferie_permesso')
                })

    logging.info(f"Numero di attività trovate: {len(attivita_list)}")
    return attivita_list


def is_date_in_range(date_str, data_inizio, data_fine):
    # Formato della data proveniente dal database Firebase
    date_format_db = "%d/%m/%Y"
    # Formato della data proveniente dal form HTML
    date_format_form = "%Y-%m-%d"

    try:
        # Converti le date di inizio e fine provenienti dal form HTML
        data_inizio_dt = datetime.strptime(data_inizio, date_format_form)
        data_fine_dt = datetime.strptime(data_fine, date_format_form)

        # Converti la data proveniente dal database Firebase
        current_date_dt = datetime.strptime(date_str, date_format_db)

        # Verifica se la data dell'attività rientra nell'intervallo di date
        return data_inizio_dt <= current_date_dt <= data_fine_dt

    except ValueError as e:
        logging.error(f"Errore nel parsing della data: {e}")
        return False



def generate_excel_contabilita(attivita_list, inizio, fine):
    file_name = f"contabilita_{inizio}_to_{fine}.xlsx"
    file_path = os.path.join(TEMP_DIR, file_name)

    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet("Contabilità")

    headers = ['Data', 'Cantiere', 'Operaio', 'Ore', 'Lavorazione']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    for row, attivita in enumerate(attivita_list, start=1):
        worksheet.write(row, 0, attivita['data'])
        worksheet.write(row, 1, attivita['cantiere'])
        worksheet.write(row, 2, attivita['operaio'])
        worksheet.write(row, 3, attivita['ore'])
        worksheet.write(row, 4, attivita['lavorazione'])

    workbook.close()
    return file_name


def generate_excel_completo(attivita_list, inizio, fine):
    file_name = f"completo_{inizio}_to_{fine}.xlsx"
    file_path = os.path.join(TEMP_DIR, file_name)

    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet("Completo")

    headers = ['Data', 'Cantiere', 'Operaio', 'Ore', 'Lavorazione', 'Pioggia_Vento', 'Ferie_Permesso']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    for row, attivita in enumerate(attivita_list, start=1):
        worksheet.write(row, 0, attivita['data'])
        worksheet.write(row, 1, attivita['cantiere'])
        worksheet.write(row, 2, attivita['operaio'])
        worksheet.write(row, 3, attivita['ore'])
        worksheet.write(row, 4, attivita['lavorazione'])
        worksheet.write(row, 5, attivita['pioggia_vento'])
        worksheet.write(row, 6, attivita['ferie_permesso'])

    workbook.close()
    return file_name


def generate_excel_buste(attivita_list, inizio, fine):
    file_name = f"buste_paga_{inizio}_to_{fine}.xlsx"
    file_path = os.path.join(TEMP_DIR, file_name)

    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet("Buste Paga")

    headers = ['Data', 'Cantiere', 'Operaio', 'Ore', 'Pioggia_Vento', 'Ferie_Permesso']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)

    for row, attivita in enumerate(attivita_list, start=1):
        worksheet.write(row, 0, attivita['data'])
        worksheet.write(row, 1, attivita['cantiere'])
        worksheet.write(row, 2, attivita['operaio'])
        worksheet.write(row, 3, attivita['ore'])
        worksheet.write(row, 4, attivita['pioggia_vento'])
        worksheet.write(row, 5, attivita['ferie_permesso'])

    workbook.close()
    return file_name


@app.route('/gestione_attivita')
def gestione_attivita():
    return render_template('gestione_attivita.html')

@app.route('/gestione_foto_bolle', methods=['GET', 'POST'])
def gestione_foto_bolle():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    
    current_year = datetime.now().year

    years = list(range(2023, current_year + 1))
    months = [f"{month:02d}" for month in range(1, 13)]
    period = ""

    if request.method == 'POST':
        start_year = request.form.get('start_year')
        start_month = request.form.get('start_month')
        end_year = request.form.get('end_year', '')
        end_month = request.form.get('end_month', '')

        logging.debug(f"Start Year: {start_year}, Start Month: {start_month}, End Year: {end_year}, End Month: {end_month}")

        action = request.form.get('action')  # Visualizza o Scarica
        
        # Imposta il periodo da visualizzare accanto a "Risultati"
        if end_year == '':
            period = f"{start_month}/{start_year}"
        else:
            period = f"da {start_month}/{start_year} a {end_month}/{end_year}"

        if action == "Visualizza Foto":
            logging.debug(f"Azione selezionata: {action}")
            photo_urls = fetch_photos(start_year, start_month, end_year, end_month)
            logging.debug(f"Numero di foto trovate: {len(photo_urls)}")
            return render_template('gestione_foto_bolle.html', photo_urls=photo_urls, years=years, months=months, period=period)
        
        elif action == "Scarica Foto":
            logging.debug(f"Azione selezionata: {action}")
            zip_filepath = create_zip(start_year, start_month, end_year, end_month)
            
            if zip_filepath is None:
                logging.debug("Nessuna foto trovata per il periodo selezionato")
                return jsonify({'error': 'Non ci sono foto per il periodo selezionato.'})
            else:
                logging.debug(f"File ZIP creato: {zip_filepath}")
                zip_url = url_for('download_zip', filename=os.path.basename(zip_filepath))
                return jsonify({'zip_url': zip_url})
            
    return render_template('gestione_foto_bolle.html', years=years, months=months, period=period)

def fetch_photos(start_year, start_month, end_year, end_month):
    start_year = int(start_year)
    start_month = int(start_month)
    photo_urls = []

    logging.debug(f"Fetching photos from {start_month}/{start_year} to {end_month}/{end_year}")

    # Se l'utente non seleziona la data di fine
    if end_year =='':
        blobs = list(bucket.list_blobs(prefix=f"DDT/{start_year}/{start_month}/"))
        logging.debug(f"Numero di blob trovati per {start_month}/{start_year}: {len(blobs)}")
        for blob in blobs:
            url = blob.generate_signed_url(timedelta(seconds=300), method='GET')
            photo_urls.append(url)
    else:
        # Logica per iterare tra più anni e mesi
        
        end_year = int(end_year)
        end_month = int(end_month)

        for year in range(start_year, end_year + 1):
            month_start = start_month if year == start_year else 1
            month_end = end_month if year == end_year else 12

            for month in range(month_start, month_end + 1):
                logging.debug(f"Fetching blobs for {month}/{year}")
                blobs = list(bucket.list_blobs(prefix=f"DDT/{year}/{month}/"))
                logging.debug(f"Numero di blob trovati per {month}/{year}: {len(blobs)}")
                for blob in blobs:
                    logging.debug(f"Blob trovato: {blob.name}")
                    url = blob.generate_signed_url(timedelta(seconds=300), method='GET')
                    photo_urls.append(url)
                    
    return photo_urls

def create_zip(start_year, start_month, end_year, end_month):
    start_year = int(start_year)
    start_month = int(start_month)

    no_photos = True  # Flag per tenere traccia se ci sono foto o meno

    if end_year == '':
        zip_filename = f"foto_{start_year}_{start_month}.zip"
        end_year = start_year
        end_month = start_month
    else:
        end_year = int(end_year)
        end_month = int(end_month)
        # Crea un file ZIP temporaneo per il range di date
        zip_filename = f"foto_{start_year}_{start_month}_to_{end_year}_{end_month}.zip"

    zip_filepath = os.path.join(TEMP_DIR, zip_filename)

    logging.debug(f"Creazione del file ZIP: {zip_filename}")

    # Apri il file ZIP
    with zipfile.ZipFile(zip_filepath, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Itera su tutti gli anni e mesi tra la data di inizio e di fine
        for year in range(start_year, end_year + 1):
            # Definisci il mese di inizio e di fine per ciascun anno
            month_start = start_month if year == start_year else 1
            month_end = end_month if year == end_year else 12

            for month in range(int(month_start), month_end + 1):
                # Ottieni i blob per l'anno e il mese correnti
                logging.debug(f"Elaborazione blobs per {month}/{year}")
                blobs = list(bucket.list_blobs(prefix=f"DDT/{year}/{month}/"))

                if len(blobs) == 0:
                    logging.debug(f"Nessun blob trovato per {month}/{year}")
                    continue  # Nessun blob per questo mese, continua con il prossimo

                no_photos = False  # Se troviamo almeno un blob, imposta no_photos a False

                for idx, blob in enumerate(blobs):
                    blob_name = blob.name
                    blob_name = '/'.join(blob_name.split('/')[1:])  # Rimuove il prefisso 'DDT'

                    logging.debug(f"Aggiunta al file ZIP: {blob_name}")

                    if blob.exists():  # Verifica che il blob esista
                        image_data = blob.download_as_bytes()  # Scarica i dati del blob
                        zip_file.writestr(f"{blob_name}", image_data)  # Salva l'immagine nel file ZIP
                    else:
                        logging.debug(f"Errore: Blob {blob_name} non trovato")

    # Se no_photos è True, significa che non ci sono foto, elimina lo zip
    if no_photos:
        logging.debug("Nessuna foto trovata, elimino il file ZIP vuoto.")
        os.remove(zip_filepath)  # Rimuovi il file ZIP che è stato creato vuoto
        return None

    logging.debug(f"File ZIP creato con successo: {zip_filepath}")
    return zip_filename

@app.route('/download_zip/<filename>')
def download_zip(filename):
    response = send_from_directory(TEMP_DIR, filename, as_attachment=True)
    return response


def clean_temp_directory():
    try:
        shutil.rmtree(TEMP_DIR)  # Rimuovi la directory temporanea e tutto il suo contenuto
        os.makedirs(TEMP_DIR, exist_ok=True)  # Ricrea la directory vuota
        logging.info("Cartella temp pulita con successo.")
    except Exception as e:
        logging.error(f"Errore durante la pulizia della cartella temp: {e}")


if __name__ == '__main__':
    app.run(debug=True)
